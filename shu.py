import cv2
import openpyxl
from datetime import datetime
import os

def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        new_sheet = wb.create_sheet(sheet_name)
        new_sheet.append(["Name", "Time"])
        return new_sheet

# Load or create the workbook
if os.path.exists("attendance.xlsx"):
    wb = openpyxl.load_workbook("attendance.xlsx")
else:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

# Create a directory to store face images
if not os.path.exists("face_images"):
    os.makedirs("face_images")

# Load the pre-trained face detection classifier
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Dictionary to store names
names = {}

# Set to keep track of marked attendance for the current day
marked_attendance = set()

# Function to mark attendance
def mark_attendance(name):
    if name in names.values() and name not in marked_attendance:
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        
        sheet = get_or_create_sheet(wb, date)
        sheet.append([name, time])
        marked_attendance.add(name)
        print(f"Attendance marked for {name} at {time}")
        
        # Update total count
        total_cell = sheet.cell(row=1, column=4)
        total_cell.value = f"Total Present: {len(marked_attendance)}"
        
        wb.save("attendance.xlsx")

# Initialize the camera
cap = cv2.VideoCapture(0)

face_count = 0

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:
        face_id = face_count
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
        
        if face_id in names:
            name = names[face_id]
            cv2.putText(frame, name, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
            mark_attendance(name)
        else:
            cv2.putText(frame, "Unknown", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

    cv2.imshow("Face Scanner Attendance", frame)

    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break
    elif key == ord('a'):
        # Add a new face
        name = input("Enter the name of the person: ")
        cv2.imwrite(f"face_images/{name}.jpg", frame)
        print(f"Face image saved for {name}")
        names[face_count] = name
        face_count += 1
        mark_attendance(name)  # Mark attendance when a new face is added

cap.release()
cv2.destroyAllWindows()

# Save the workbook
wb.save("attendance.xlsx")

# Print attendance for the current day
current_date = datetime.now().strftime("%Y-%m-%d")
if current_date in wb.sheetnames:
    print(f"\nAttendance Record for {current_date}:")
    for row in wb[current_date].iter_rows(values_only=True):
        print(row)
    print(f"Total Present: {len(marked_attendance)}")
else:
    print("No attendance recorded for today.")