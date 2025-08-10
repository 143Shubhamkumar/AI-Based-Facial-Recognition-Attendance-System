import cv2
import numpy as np
import openpyxl
from datetime import datetime
import os
import pickle

def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        new_sheet = wb.create_sheet(sheet_name)
        new_sheet.append(["Name", "Time"])
        return new_sheet

# Initialize variables
known_face_features = []
known_face_names = []
name_to_label = {}
current_label = 0

# Load or create the workbook
if os.path.exists("attendance.xlsx"):
    wb = openpyxl.load_workbook("attendance.xlsx")
else:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

# Create directories to store face images and features
if not os.path.exists("face_images"):
    os.makedirs("face_images")
if not os.path.exists("face_features"):
    os.makedirs("face_features")

# Load existing face features if available
if os.path.exists("face_features/features.pkl"):
    with open("face_features/features.pkl", "rb") as f:
        loaded_data = pickle.load(f)
        if isinstance(loaded_data, tuple) and len(loaded_data) == 2:
            known_face_features, known_face_names = loaded_data
            # Reconstruct name_to_label and current_label
            name_to_label = {name: i for i, name in enumerate(known_face_names)}
            current_label = len(known_face_names)
        elif isinstance(loaded_data, tuple) and len(loaded_data) == 4:
            known_face_features, known_face_names, name_to_label, current_label = loaded_data
        else:
            print("Unexpected data format in features.pkl. Initializing with empty data.")

# Set to keep track of marked attendance for the current day
marked_attendance = set()

# Function to mark attendance
def mark_attendance(name):
    if name not in marked_attendance:
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        
        sheet = get_or_create_sheet(wb, date)
        sheet.append([name, time])
        marked_attendance.add(name)
        print(f"Attendance marked for {name} at {time}")
        
        # Update total count
        total_cell = sheet.cell(row=1, column=4)
        total_cell.value = f"Total Present: {len(marked_attendance)}"
        
        wb.save("attendance.xlsx")

# Initialize face detector and feature extractor
face_detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
face_recognizer = cv2.face.LBPHFaceRecognizer_create()

if known_face_features:
    labels = [name_to_label[name] for name in known_face_names]
    face_recognizer.train(known_face_features, np.array(labels))

# Initialize the camera
cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_detector.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:
        face_roi = gray[y:y+h, x:x+w]
        
        if known_face_features:
            label, confidence = face_recognizer.predict(face_roi)
            if confidence < 100:  # You may need to adjust this threshold
                name = known_face_names[label]
                mark_attendance(name)
            else:
                name = "Unknown"
        else:
            name = "Unknown"

        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
        cv2.putText(frame, name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

    cv2.imshow("Face Scanner Attendance", frame)

    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break
    elif key == ord('a'):
        # Add a new face
        name = input("Enter the name of the person: ")
        cv2.imwrite(f"face_images/{name}.jpg", frame)
        print(f"Face image saved for {name}")
        
        # Compute and save face features
        face_roi = gray[y:y+h, x:x+w]
        known_face_features.append(face_roi)
        known_face_names.append(name)

        if name not in name_to_label:
            name_to_label[name] = current_label
            current_label += 1
        
        # Retrain the recognizer
        labels = [name_to_label[name] for name in known_face_names]
        face_recognizer.train(known_face_features, np.array(labels))
        
        # Save updated features
        with open("face_features/features.pkl", "wb") as f:
            pickle.dump((known_face_features, known_face_names, name_to_label, current_label), f)
        
        mark_attendance(name)

cap.release()
cv2.destroyAllWindows()

# Save the workbook
wb.save("attendance.xlsx")

# Print attendance for the current day
current_date = datetime.now().strftime("%Y-%m-%d")
if current_date in wb.sheetnames:
    print(f"\nAttendance Record for {current_date}:")
    for row in wb[current_date].iter_rows(values_only=True):
        print(row)
    print(f"Total Present: {len(marked_attendance)}")
else:
    print("No attendance recorded for today.")